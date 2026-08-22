#!/usr/bin/env python3
"""
読み（カナ）をキーに、日本古典対照分類語彙表.xlsx の2シートから
`対照語彙表`の候補を検索する（読み取り専用）。

tsj_wakun_ref_candidates.py は entry_text（見出し漢字）の完全一致でしか
探せないため、「ヤマノミネ（嶼）」のような定義訓・複合的な読みに対しては
ほぼ候補が見つからない。本スクリプトは、そこから抽出した語幹・見出し語
（例: 「ミネ」「マク」「ツク」）を直接検索するための道具。
語幹の抽出（複合語の分解、動詞の終止形への還元）自体は行わない
——これは読み手の言語知識が必要な作業であり、本スクリプトはあくまで
「抽出済みの読みで参照表を引く」検索窓の役割に徹する。

使い方:
  # ひらがな・カタカナどちらでもよい。複数指定すると1回の起動でまとめて引ける
  # （xlsx読み込みに数秒かかるため、まとめて渡す方が効率的）
  python3 samples/scripts/taisho_reading_lookup.py みね まく つく

  # 完全一致で見つからない場合に部分一致（読みの前方一致・部分一致）も試す
  python3 samples/scripts/taisho_reading_lookup.py --fuzzy へしくき
"""

from __future__ import annotations

import argparse
import sys
import unicodedata
from pathlib import Path

import openpyxl

DEFAULT_CLASSICAL_XLSX = Path("xlsx/日本古典対照分類語彙表.xlsx")


def to_hiragana(text: str) -> str:
    return "".join(chr(ord(c) - 0x60) if "ァ" <= c <= "ヶ" else c for c in text)


def normalize(text: str) -> str:
    return to_hiragana(unicodedata.normalize("NFKC", text)).strip()


def load_classical(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["日本古典対照分類語彙表"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {h: i for i, h in enumerate(header)}
    out = []
    for r in rows:
        midashi = r[idx["見出し"]]
        if not midashi:
            continue
        out.append(
            {
                "reading": normalize(midashi),
                "kanji": r[idx["漢字"]] or "",
                "pos_raw": r[idx["品詞"]] or "",
                "word_type": r[idx["語種"]] or "",
                "category": r[idx["意味分類"]] or "",
                "source": "classical",
            }
        )
    return out


def load_modern(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["bunruidb-fam"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {h: i for i, h in enumerate(header)}
    out = []
    for r in rows:
        yomi = r[idx["読み"]]
        if not yomi:
            continue
        bangou = r[idx["分類番号"]]
        koumoku = r[idx["分類項目"]]
        out.append(
            {
                "reading": normalize(yomi),
                "kanji": r[idx["見出し本体"]] or "",
                "pos_raw": r[idx["類"]] or "",
                "word_type": r[idx["部門"]] or "",
                "category": f"{bangou}({koumoku})" if bangou else "",
                "source": "modern",
            }
        )
    return out


def search(entries: list[dict], query: str, fuzzy: bool) -> list[dict]:
    exact = [e for e in entries if e["reading"] == query]
    if exact or not fuzzy:
        return exact
    return [e for e in entries if query in e["reading"] or e["reading"] in query]


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("queries", nargs="+", help="検索する読み（ひらがな・カタカナ可）")
    parser.add_argument("--classical-xlsx", type=Path, default=DEFAULT_CLASSICAL_XLSX)
    parser.add_argument("--fuzzy", action="store_true", help="完全一致がない場合に部分一致も試す")
    parser.add_argument("--max-results", type=int, default=20)
    args = parser.parse_args()

    print(f"Loading {args.classical_xlsx} ...", file=sys.stderr)
    classical = load_classical(args.classical_xlsx)
    modern = load_modern(args.classical_xlsx)
    print(f"Loaded: classical={len(classical)} modern={len(modern)}", file=sys.stderr)

    for raw_query in args.queries:
        query = normalize(raw_query)
        print(f"\n=== query: {raw_query!r} (normalized: {query!r}) ===")
        results = search(classical, query, args.fuzzy) + search(modern, query, args.fuzzy)
        if not results:
            print("  (no match)")
            continue
        for r in results[: args.max_results]:
            print(
                f"  [{r['source']}] reading={r['reading']!r} kanji={r['kanji']!r} "
                f"pos_raw={r['pos_raw']!r} category={r['category']!r}"
            )
        if len(results) > args.max_results:
            print(f"  ... and {len(results) - args.max_results} more (raise --max-results to see all)")


if __name__ == "__main__":
    main()
