#!/usr/bin/env python3
"""
tsj_wakun*.xlsx「和訓改訂版」シートの `対照語彙表`・`pos` 列が未記入の行について、
日本古典対照分類語彙表.xlsx の2シート（「日本古典対照分類語彙表」＝古典語彙、
「bunruidb-fam」＝現代語彙）を漢字（見出し）で突合し、候補を提示する
（読み取り専用。どちらのxlsxも書き換えない）。

候補は自動で確定しない。「日本古典対照分類語彙表」の `漢字` 列と完全一致する行を
優先候補、見つからなければ「bunruidb-fam」の `見出し本体` 列と完全一致する行を
次点候補として提示する。1つの漢字に複数の見出し（同綴異語）がある場合は、
`reading_kana_kanji`/`reading_historical_kana` との読みの一致も試み、一致すれば
その候補を先頭に出す。

日本古典対照分類語彙表の `品詞` は活用型ベースの分類（動四・動下二・形動…）で
あり、tsj_wakun側の `pos`（名詞・動詞・連語（…）…）とは語彙が異なる。
このスクリプトは変換表で機械的に対応付けたりはしない。品詞候補はあくまで
「古典語彙表側の生の値」をそのまま提示するので、`pos` 列への転記時は
tsj_wakun側の語彙に合わせて人が判断すること。

使い方:
  python3 samples/scripts/tsj_wakun_ref_candidates.py \\
      xlsx/tsj_wakun20260819.xlsx xlsx/日本古典対照分類語彙表.xlsx

  # 対象列を絞る（デフォルトは両方）
  python3 samples/scripts/tsj_wakun_ref_candidates.py \\
      xlsx/tsj_wakun20260819.xlsx xlsx/日本古典対照分類語彙表.xlsx --fields pos

  # レポートをファイルへ
  python3 samples/scripts/tsj_wakun_ref_candidates.py \\
      xlsx/tsj_wakun20260819.xlsx xlsx/日本古典対照分類語彙表.xlsx --output report.tsv --format tsv
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl

FIELD_CHOICES = ["対照語彙表", "pos"]


def load_sheet_rows(path: Path, sheet_name: str) -> tuple[list[str], list[tuple]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet {sheet_name!r} not found in {path}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    rows = [r for r in rows_iter if any(c is not None and str(c) != "" for c in r)]
    return header, rows


def strip_reading_paren(reading: str | None) -> str:
    """'フスマ（衾）' のようなカナ表記から括弧部分を除いた読みだけを返す。"""
    if not reading:
        return ""
    return re.sub(r"[（(].*?[）)]\s*$", "", reading).strip()


def build_classical_index(rows: list[tuple], header: list[str]) -> dict[str, list[dict]]:
    idx = {h: i for i, h in enumerate(header)}
    kanji_i = idx["漢字"]
    kana_i = idx["仮名（漢字）"]
    hinshi_i = idx["品詞"]
    gosyu_i = idx["語種"]
    imibunrui_i = idx["意味分類"]
    index: dict[str, list[dict]] = {}
    for r in rows:
        kanji = r[kanji_i]
        if not kanji:
            continue
        index.setdefault(kanji, []).append(
            {
                "reading": strip_reading_paren(r[kana_i]),
                "pos_raw": r[hinshi_i] or "",
                "word_type": r[gosyu_i] or "",
                "category": r[imibunrui_i] or "",
            }
        )
    return index


def build_modern_index(rows: list[tuple], header: list[str]) -> dict[str, list[dict]]:
    idx = {h: i for i, h in enumerate(header)}
    honta_i = idx["見出し本体"]
    yomi_i = idx["読み"]
    ruib_i = idx["類"]
    bunrui_koumoku_i = idx["分類項目"]
    bunrui_bangou_i = idx["分類番号"]
    index: dict[str, list[dict]] = {}
    for r in rows:
        honta = r[honta_i]
        if not honta:
            continue
        index.setdefault(honta, []).append(
            {
                "reading": r[yomi_i] or "",
                "pos_raw": r[ruib_i] or "",
                "category": f"{r[bunrui_bangou_i]}({r[bunrui_koumoku_i]})" if r[bunrui_bangou_i] else "",
            }
        )
    return index


def rank_candidates(candidates: list[dict], target_readings: set[str]) -> list[dict]:
    def score(c: dict) -> int:
        return 0 if c["reading"] in target_readings else 1

    return sorted(candidates, key=score)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("tsj_wakun_xlsx", type=Path)
    parser.add_argument("classical_xlsx", type=Path)
    parser.add_argument("--tsj-sheet", default="和訓改訂版")
    parser.add_argument("--classical-sheet", default="日本古典対照分類語彙表")
    parser.add_argument("--modern-sheet", default="bunruidb-fam")
    parser.add_argument("--fields", default=",".join(FIELD_CHOICES))
    parser.add_argument("--max-candidates", type=int, default=3)
    parser.add_argument("--format", choices=["text", "tsv"], default="text")
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    fields = [f.strip() for f in args.fields.split(",") if f.strip()]
    for f in fields:
        if f not in FIELD_CHOICES:
            raise SystemExit(f"Unknown field {f!r}; choose from {FIELD_CHOICES}")

    tsj_header, tsj_rows = load_sheet_rows(args.tsj_wakun_xlsx, args.tsj_sheet)
    classical_header, classical_rows = load_sheet_rows(args.classical_xlsx, args.classical_sheet)
    modern_header, modern_rows = load_sheet_rows(args.classical_xlsx, args.modern_sheet)

    classical_index = build_classical_index(classical_rows, classical_header)
    modern_index = build_modern_index(modern_rows, modern_header)

    tidx = {h: i for i, h in enumerate(tsj_header)}
    entry_i = tidx["entry_text"]
    sjw_i = tidx["sj_w_id"]
    rkk_i = tidx["reading_kana_kanji"]
    rhk_i = tidx["reading_historical_kana"]
    pos_i = tidx["pos"]
    taisho_i = tidx["対照語彙表"]

    results = []
    for r in tsj_rows:
        missing = []
        if "対照語彙表" in fields and r[taisho_i] in (None, ""):
            missing.append("対照語彙表")
        if "pos" in fields and r[pos_i] in (None, ""):
            missing.append("pos")
        if not missing:
            continue

        entry_text = r[entry_i]
        target_readings = {strip_reading_paren(r[rkk_i]), r[rhk_i] or ""}
        target_readings.discard("")

        classical_matches = rank_candidates(list(classical_index.get(entry_text, [])), target_readings)
        modern_matches = rank_candidates(list(modern_index.get(entry_text, [])), target_readings)

        results.append(
            {
                "sj_w_id": r[sjw_i],
                "entry_text": entry_text,
                "reading": r[rkk_i] or "",
                "missing": missing,
                "classical_matches": classical_matches[: args.max_candidates],
                "modern_matches": [] if classical_matches else modern_matches[: args.max_candidates],
            }
        )

    render(results, args.format, args.output, len(tsj_rows))


def render(results: list[dict], fmt: str, output: Path | None, total_rows: int) -> None:
    lines = []
    no_match = sum(1 for r in results if not r["classical_matches"] and not r["modern_matches"])
    lines.append(f"tsj_wakun rows: {total_rows}")
    lines.append(f"rows needing candidates: {len(results)}")
    lines.append(f"rows with no candidate found in either reference sheet: {no_match}")
    lines.append("")

    if fmt == "tsv":
        lines = ["sj_w_id\tentry_text\treading\tmissing\tsource\tcandidate_reading\tpos_raw\tcategory"]
        for r in results:
            matches = [("classical", m) for m in r["classical_matches"]] + [
                ("modern", m) for m in r["modern_matches"]
            ]
            if not matches:
                lines.append(f"{r['sj_w_id']}\t{r['entry_text']}\t{r['reading']}\t{','.join(r['missing'])}\t(no match)\t\t\t")
                continue
            for source, m in matches:
                lines.append(
                    f"{r['sj_w_id']}\t{r['entry_text']}\t{r['reading']}\t{','.join(r['missing'])}\t"
                    f"{source}\t{m['reading']}\t{m['pos_raw']}\t{m['category']}"
                )
    else:
        for r in results:
            lines.append(f"[{r['sj_w_id']}] {r['entry_text']} ({r['reading']}) missing={','.join(r['missing'])}")
            matches = [("classical", m) for m in r["classical_matches"]] + [
                ("modern", m) for m in r["modern_matches"]
            ]
            if not matches:
                lines.append("  (no candidate found)")
            for source, m in matches:
                lines.append(f"  [{source}] reading={m['reading']!r} pos_raw={m['pos_raw']!r} category={m['category']!r}")

    text = "\n".join(lines)
    if output is None:
        print(text)
    else:
        output.write_text(text + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
