#!/usr/bin/env python3
"""
Excel（xlsx）の改訂シートとHDICのTSVを、共通のキー列で突き合わせて
列ごとの差分を報告する（デフォルトでは読み取り専用。データは変更しない）。

実例: xlsx/tsj_wakun20260625.xlsx の「和訓改訂版」シートと TSJ_wakun.tsv を
sj_w_id で突合し、pos列・remarks列だけ改訂を反映した（コミット 7f98580）。
このときentry_text等の列はxlsx側が旧版だったため意図的に据え置かれた。
「どの列をxlsx側で上書きするか」は資料批判を要する編集判断であり、本スクリプトは
その判断を自動で下さない。差分を列ごとに報告するところまでを行い、適用する列は
必ず人（ユーザー）に選んでもらう。

使い方:
  # 差分レポートのみ（読み取り専用）
  python3 samples/scripts/xlsx_tsv_diff.py \\
      xlsx/tsj_wakun20260625.xlsx "和訓改訂版" TSJ_wakun.tsv sj_w_id

  # 確認の上、承認された列だけをTSVに適用する
  python3 samples/scripts/xlsx_tsv_diff.py \\
      xlsx/tsj_wakun20260625.xlsx "和訓改訂版" TSJ_wakun.tsv sj_w_id \\
      --apply-columns pos,remarks

--apply-columns で指定した列のみ、キーが一致する行のデータセルを書き換える。
ヘッダーのコメント行・列構成・xlsx側にしか存在しないキーの行・承認されていない
列は一切変更しない。書き換え後はcount_basic_stats.py / validate_hdic_integrity.py /
finalize_hdic_edit.py を別途実行すること（本スクリプトはそれらを呼ばない）。
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

import openpyxl


def read_tsv(path: Path) -> tuple[list[str], list[list[str]], list[str]]:
    """(header, data_rows, all_lines) を返す。all_linesは書き戻し用の生の行。"""
    header: list[str] | None = None
    data_rows: list[list[str]] = []
    all_lines = path.read_text(encoding="utf-8-sig").splitlines(keepends=True)
    for line in all_lines:
        stripped = line.rstrip("\n")
        row = next(csv.reader([stripped], delimiter="\t"))
        if not row or all(cell == "" for cell in row):
            continue
        if row[0].lstrip("﻿").startswith("#"):
            continue
        if header is None:
            header = row
            continue
        data_rows.append(row)
    if header is None:
        raise SystemExit(f"No header row found in {path}")
    return header, data_rows, all_lines


def read_xlsx_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[list]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet {sheet_name!r} not found in {path}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    rows = [list(r) for r in rows_iter if any(c is not None and str(c) != "" for c in r)]
    return header, rows


def cell_str(value) -> str:
    if value is None:
        return ""
    return str(value)


def build_diff(
    tsv_header: list[str],
    tsv_rows: list[list[str]],
    xlsx_header: list[str],
    xlsx_rows: list[list],
    key_column: str,
) -> dict:
    if key_column not in tsv_header:
        raise SystemExit(f"Key column {key_column!r} not found in TSV header: {tsv_header}")
    if key_column not in xlsx_header:
        raise SystemExit(f"Key column {key_column!r} not found in xlsx header: {xlsx_header}")

    tsv_key_idx = tsv_header.index(key_column)
    xlsx_key_idx = xlsx_header.index(key_column)

    tsv_by_key = {row[tsv_key_idx]: row for row in tsv_rows if tsv_key_idx < len(row) and row[tsv_key_idx]}
    xlsx_by_key = {
        cell_str(row[xlsx_key_idx]): row for row in xlsx_rows if xlsx_key_idx < len(row) and row[xlsx_key_idx]
    }

    common_columns = [c for c in tsv_header if c in xlsx_header and c != key_column]
    xlsx_only_columns = [c for c in xlsx_header if c not in tsv_header]
    tsv_only_columns = [c for c in tsv_header if c not in xlsx_header]

    only_in_xlsx = sorted(set(xlsx_by_key) - set(tsv_by_key))
    only_in_tsv = sorted(set(tsv_by_key) - set(xlsx_by_key))

    per_column_diffs: dict[str, list[dict]] = {c: [] for c in common_columns}
    for key, xlsx_row in xlsx_by_key.items():
        tsv_row = tsv_by_key.get(key)
        if tsv_row is None:
            continue
        for col in common_columns:
            t_idx = tsv_header.index(col)
            x_idx = xlsx_header.index(col)
            tsv_val = tsv_row[t_idx] if t_idx < len(tsv_row) else ""
            xlsx_val = cell_str(xlsx_row[x_idx]) if x_idx < len(xlsx_row) else ""
            if tsv_val != xlsx_val:
                per_column_diffs[col].append({"key": key, "tsv": tsv_val, "xlsx": xlsx_val})

    return {
        "common_columns": common_columns,
        "xlsx_only_columns": xlsx_only_columns,
        "tsv_only_columns": tsv_only_columns,
        "only_in_xlsx_keys": only_in_xlsx,
        "only_in_tsv_keys": only_in_tsv,
        "per_column_diffs": per_column_diffs,
    }


def render_report(diff: dict) -> str:
    lines = ["xlsx / TSV diff report", ""]
    lines.append(f"columns compared: {', '.join(diff['common_columns'])}")
    if diff["xlsx_only_columns"]:
        lines.append(f"columns only in xlsx (not compared): {', '.join(diff['xlsx_only_columns'])}")
    if diff["tsv_only_columns"]:
        lines.append(f"columns only in TSV (not compared): {', '.join(diff['tsv_only_columns'])}")
    lines.append(f"keys only in xlsx (no matching TSV row): {len(diff['only_in_xlsx_keys'])}")
    if diff["only_in_xlsx_keys"]:
        lines.append(f"  {diff['only_in_xlsx_keys'][:10]}")
    lines.append(f"keys only in TSV (no matching xlsx row): {len(diff['only_in_tsv_keys'])}")
    if diff["only_in_tsv_keys"]:
        lines.append(f"  {diff['only_in_tsv_keys'][:10]}")
    lines.append("")
    total = 0
    for col, changes in diff["per_column_diffs"].items():
        if not changes:
            continue
        total += len(changes)
        lines.append(f"[{col}] {len(changes)} row(s) differ")
        for c in changes[:5]:
            lines.append(f"  {c['key']}: tsv={c['tsv']!r} -> xlsx={c['xlsx']!r}")
        if len(changes) > 5:
            lines.append(f"  ... and {len(changes) - 5} more")
    if total == 0:
        lines.append("No cell differences in common columns.")
    return "\n".join(lines)


def apply_columns(
    tsv_path: Path,
    tsv_header: list[str],
    all_lines: list[str],
    diff: dict,
    columns: list[str],
) -> int:
    changes_by_key: dict[str, dict[str, str]] = {}
    for col in columns:
        if col not in diff["per_column_diffs"]:
            raise SystemExit(f"Column {col!r} was not compared (not common to both sources)")
        for c in diff["per_column_diffs"][col]:
            changes_by_key.setdefault(c["key"], {})[col] = c["xlsx"]

    key_column_idx = tsv_header.index(diff["_key_column"])
    applied = 0
    header_seen = False
    for i, line in enumerate(all_lines):
        stripped = line.rstrip("\n")
        row = next(csv.reader([stripped], delimiter="\t"))
        if not row or all(cell == "" for cell in row):
            continue
        if row[0].lstrip("﻿").startswith("#"):
            continue
        if not header_seen:
            header_seen = True
            continue
        key = row[key_column_idx] if key_column_idx < len(row) else ""
        if key not in changes_by_key:
            continue
        new_row = list(row)
        while len(new_row) < len(tsv_header):
            new_row.append("")
        for col, new_val in changes_by_key[key].items():
            new_row[tsv_header.index(col)] = new_val
        all_lines[i] = "\t".join(new_row) + "\n"
        applied += 1

    tsv_path.write_text("".join(all_lines), encoding="utf-8")
    return applied


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("sheet_name")
    parser.add_argument("tsv_path", type=Path)
    parser.add_argument("key_column")
    parser.add_argument("--format", choices=["text", "json"], default="text")
    parser.add_argument(
        "--apply-columns",
        help="Comma-separated column names to write into the TSV (only rows where the value differs)",
    )
    args = parser.parse_args()

    tsv_header, tsv_rows, all_lines = read_tsv(args.tsv_path)
    xlsx_header, xlsx_rows = read_xlsx_sheet(args.xlsx_path, args.sheet_name)
    diff = build_diff(tsv_header, tsv_rows, xlsx_header, xlsx_rows, args.key_column)
    diff["_key_column"] = args.key_column

    if args.apply_columns:
        columns = [c.strip() for c in args.apply_columns.split(",") if c.strip()]
        applied = apply_columns(args.tsv_path, tsv_header, all_lines, diff, columns)
        print(f"Applied changes to {applied} row(s) for column(s): {', '.join(columns)}")
        print("Run count_basic_stats.py / validate_hdic_integrity.py / finalize_hdic_edit.py next.")
        return

    diff.pop("_key_column", None)
    if args.format == "json":
        print(json.dumps(diff, ensure_ascii=False, indent=2))
    else:
        print(render_report(diff))

    total_diffs = sum(len(v) for v in diff["per_column_diffs"].values())
    sys.exit(1 if total_diffs else 0)


if __name__ == "__main__":
    main()
